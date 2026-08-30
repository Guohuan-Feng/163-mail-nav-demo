#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
163 邮箱基金净值附件提取（演示版）

范围：
    - 手动执行；
    - 只读取 163 邮箱 INBOX 中的 .xlsx 附件；
    - 识别常见净值表头，输出固定 10 列 Excel；
    - 不发邮件、不创建计划任务、不修改邮箱已读状态。

真实邮箱模式：
    python nav_mail_demo.py

离线演示模式（不会连接邮箱）：
    python nav_mail_demo.py --input-dir sample_input
"""

from __future__ import annotations

import argparse
import email
import imaplib
import math
import re
import sys
import traceback
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from email.header import decode_header
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


APP_DIR = Path(__file__).resolve().parent

# 在 Windows 的非 UTF-8 终端中也保证中文日志不会让程序中断。
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, OSError):
        pass

OUTPUT_COLUMNS = (
    "基金代码",
    "基金名称",
    "产品代码",
    "产品名称",
    "净值日期",
    "单位净值",
    "累计净值",
    "来源附件",
    "来源工作表",
    "来源行号",
)

# 为了让演示版行为可解释，表头匹配只认下面这些常见名称。
HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "基金代码": ("基金代码", "基金编号", "fund code", "fund_code"),
    "基金名称": ("基金名称", "基金简称", "fund name", "fund_name"),
    "产品代码": ("产品代码", "产品编号", "product code", "product_code"),
    "产品名称": ("产品名称", "产品全称", "产品简称", "product name", "product_name"),
    "净值日期": ("净值日期", "估值日期", "公布日期", "交易日期", "日期", "value date", "date"),
    "单位净值": (
        "单位净值",
        "每份净值",
        "净值",
        "unit nav",
        "unit_nav",
        "nav",
    ),
    "累计净值": (
        "累计单位净值",
        "累计净值",
        "累积净值",
        "累计",
        "acc nav",
        "acc_nav",
        "cum nav",
    ),
}


class DemoError(RuntimeError):
    """面向使用者的、可读的异常。"""


@dataclass(frozen=True)
class Attachment:
    path: Path
    original_name: str
    source_mail: str


@dataclass(frozen=True)
class ExtractedRow:
    values: tuple[Any, ...]
    dedupe_key: tuple[str, ...]


def info(message: str) -> None:
    print(f"[信息] {message}")


def ok(message: str) -> None:
    print(f"[完成] {message}")


def warn(message: str) -> None:
    print(f"[跳过] {message}")


def normalize_header(value: Any) -> str:
    """把表头统一成便于比对的形式。"""
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"[\s_\-()（）\[\]【】:/：.]+", "", text)
    return text


_ALIAS_INDEX: dict[str, str] = {}
for _target, _aliases in HEADER_ALIASES.items():
    for _alias in _aliases:
        _ALIAS_INDEX[normalize_header(_alias)] = _target


def canonical_header(value: Any) -> str | None:
    """返回固定字段名；对更具体的累计净值优先匹配。"""
    normalized = normalize_header(value)
    if not normalized:
        return None
    exact = _ALIAS_INDEX.get(normalized)
    if exact:
        return exact
    # “累计单位净值”不能被“单位净值”提前抢走，因此按别名长度倒序。
    for alias, target in sorted(_ALIAS_INDEX.items(), key=lambda item: len(item[0]), reverse=True):
        if len(alias) >= 3 and alias in normalized:
            return target
    return None


def text_value(value: Any) -> str:
    """将单元格转为去空白的文本，不把 nan 写入结果。"""
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none", "nat"} else text


def code_value(value: Any) -> str:
    """代码列保持文本；Excel 已丢失的前导零无法在此恢复。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
    text = text_value(value)
    return re.sub(r"\.0$", "", text)


def parse_date(value: Any) -> date | None:
    """处理日期对象、Excel 序列号和常见中文/英文日期文本。"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # 合理范围内的 Excel 日期序列号。
        if 20_000 <= float(value) <= 70_000:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    raw = text_value(value)
    if not raw:
        return None
    normalized = raw.replace("年", "-").replace("月", "-").replace("日", "")
    normalized = normalized.replace(".", "-").replace("/", "-")
    for fmt in ("%Y-%m-%d", "%Y%m%d", "%d-%m-%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(normalized, fmt).date()
        except ValueError:
            pass
    return None


def parse_number(value: Any) -> float | None:
    """将净值转换为数值；空值保留为空。"""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    raw = text_value(value)
    if not raw:
        return None
    normalized = raw.replace(",", "").replace("，", "").replace("元", "").strip()
    try:
        return float(normalized)
    except ValueError:
        return None


def load_env_file(path: Path) -> dict[str, str]:
    """读取简单 KEY=VALUE 配置，不输出任何密钥。"""
    if not path.exists():
        raise DemoError(
            f"未找到配置文件：{path.name}。请把 config.example.env 复制为 config.env 后再填写 163 邮箱授权码。"
        )
    result: dict[str, str] = {}
    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key, value = key.strip(), value.strip()
        if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
            value = value[1:-1]
        result[key] = value
    return result


def config_path_value(value: str, default: str) -> Path:
    path = Path(value or default)
    return path if path.is_absolute() else APP_DIR / path


def decode_mime_text(raw: str | None) -> str:
    if not raw:
        return "（无主题）"
    chunks: list[str] = []
    for part, charset in decode_header(raw):
        if isinstance(part, bytes):
            for encoding in (charset, "utf-8", "gb18030"):
                try:
                    chunks.append(part.decode(encoding or "utf-8", errors="replace"))
                    break
                except (LookupError, UnicodeError):
                    continue
        else:
            chunks.append(part)
    return "".join(chunks).strip() or "（无主题）"


def safe_filename(name: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", Path(name).name).strip(" .")
    return cleaned or "attachment.xlsx"


def unique_path(directory: Path, filename: str) -> Path:
    candidate = directory / filename
    counter = 2
    while candidate.exists():
        candidate = directory / f"{Path(filename).stem}_{counter}{Path(filename).suffix}"
        counter += 1
    return candidate


def fetch_from_163(
    config: dict[str, str],
    run_directory: Path,
    days: int,
    subject_keyword: str,
    max_attachment_mb: int,
) -> list[Attachment]:
    """以只读方式抓取 163 INBOX 的 .xlsx 附件。"""
    user = config.get("IMAP_USER", "").strip()
    auth_code = config.get("IMAP_AUTH_CODE", "").strip()
    host = config.get("IMAP_HOST", "imap.163.com").strip() or "imap.163.com"
    try:
        port = int(config.get("IMAP_PORT", "993"))
    except ValueError as exc:
        raise DemoError("IMAP_PORT 必须是数字。") from exc

    if not user or user.startswith("your_"):
        raise DemoError("请在 config.env 填写 IMAP_USER。")
    if not auth_code or auth_code.startswith("replace_"):
        raise DemoError("请在 config.env 填写 IMAP_AUTH_CODE（163 的 IMAP 授权码，不是网页登录密码）。")
    if not user.lower().endswith("@163.com"):
        raise DemoError("此演示版限定 163 邮箱，请确认 IMAP_USER 以 @163.com 结尾。")

    run_directory.mkdir(parents=True, exist_ok=True)
    since_date = date.today() - timedelta(days=max(days, 0))
    since_term = since_date.strftime("%d-%b-%Y")
    attachments: list[Attachment] = []
    client: imaplib.IMAP4_SSL | None = None

    try:
        info(f"正在以只读方式连接 {host}:{port} …")
        client = imaplib.IMAP4_SSL(host, port)
        client.login(user, auth_code)
        status, _ = client.select("INBOX", readonly=True)
        if status != "OK":
            raise DemoError("无法以只读方式打开 INBOX。")
        status, response = client.search(None, "SINCE", since_term)
        if status != "OK":
            raise DemoError("邮箱搜索失败。")
        mail_ids = list(reversed((response[0] or b"").split()))
        info(f"检索到 {len(mail_ids)} 封最近 {days} 天的邮件。")

        for mail_id in mail_ids:
            status, payload = client.fetch(mail_id, "(RFC822)")
            if status != "OK":
                warn(f"邮件 ID {mail_id.decode(errors='ignore')} 读取失败。")
                continue
            raw_mail = next(
                (
                    item[1]
                    for item in payload
                    if isinstance(item, tuple) and len(item) > 1 and isinstance(item[1], bytes)
                ),
                None,
            )
            if raw_mail is None:
                continue
            message = email.message_from_bytes(raw_mail)
            subject = decode_mime_text(message.get("Subject"))
            if subject_keyword and subject_keyword.lower() not in subject.lower():
                continue

            attachment_no = 0
            for part in message.walk():
                if part.is_multipart():
                    continue
                filename = part.get_filename()
                if not filename:
                    continue
                decoded_filename = decode_mime_text(filename)
                if Path(decoded_filename).suffix.lower() != ".xlsx":
                    warn(f"仅支持 .xlsx，已跳过：{decoded_filename}")
                    continue
                data = part.get_payload(decode=True)
                if not data:
                    warn(f"附件为空，已跳过：{decoded_filename}")
                    continue
                if len(data) > max_attachment_mb * 1024 * 1024:
                    warn(f"附件超过 {max_attachment_mb}MB，已跳过：{decoded_filename}")
                    continue
                attachment_no += 1
                local_name = f"{mail_id.decode(errors='ignore')}_{attachment_no}_{safe_filename(decoded_filename)}"
                target = unique_path(run_directory, local_name)
                target.write_bytes(data)
                attachments.append(
                    Attachment(path=target, original_name=decoded_filename, source_mail=subject)
                )
                ok(f"已下载附件：{decoded_filename}（主题：{subject}）")
    except imaplib.IMAP4.error as exc:
        raise DemoError(
            "163 IMAP 登录失败。请确认已开启 IMAP 服务，并在 config.env 填写新的 IMAP 授权码；"
            "不要填写网页登录密码。"
        ) from exc
    except OSError as exc:
        raise DemoError(f"无法连接 163 IMAP 服务器：{exc}") from exc
    finally:
        if client is not None:
            try:
                client.logout()
            except Exception:
                pass
    return attachments


def attachments_from_local(input_dir: Path) -> list[Attachment]:
    """离线模式：读取本地目录，不加载配置，也不会连接邮箱。"""
    if not input_dir.is_dir():
        raise DemoError(f"--input-dir 不是有效目录：{input_dir}")
    files = sorted(path for path in input_dir.iterdir() if path.is_file())
    xlsx_files = [path for path in files if path.suffix.lower() == ".xlsx"]
    for path in files:
        if path.suffix.lower() == ".xls":
            warn(f"仅支持 .xlsx，已跳过：{path.name}")
    if not xlsx_files:
        raise DemoError(f"目录中没有 .xlsx 文件：{input_dir}")
    info(f"离线模式：发现 {len(xlsx_files)} 个 .xlsx 文件，不会连接邮箱。")
    return [Attachment(path=path, original_name=path.name, source_mail="本地离线测试") for path in xlsx_files]


def find_header(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]] | None:
    """在前 20 行找字段最多、且含日期和单位净值的表头。"""
    best: tuple[int, dict[str, int]] | None = None
    for index, row in enumerate(rows[:20]):
        mapping: dict[str, int] = {}
        for column, value in enumerate(row):
            target = canonical_header(value)
            if target and target not in mapping:
                mapping[target] = column
        required = {"净值日期", "单位净值"}
        if required.issubset(mapping) and len(mapping) >= 3:
            if best is None or len(mapping) > len(best[1]):
                best = (index, mapping)
    return best


def cell_at(row: tuple[Any, ...], column: int | None) -> Any:
    if column is None or column >= len(row):
        return None
    return row[column]


def extract_rows_from_attachment(attachment: Attachment) -> tuple[list[ExtractedRow], list[str]]:
    """读取所有工作表，提取符合固定字段的记录。"""
    extracted: list[ExtractedRow] = []
    issues: list[str] = []
    try:
        workbook = load_workbook(attachment.path, read_only=True, data_only=True)
    except Exception as exc:
        return [], [f"{attachment.original_name} 无法打开：{exc}"]

    try:
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            matched = find_header(rows)
            if not matched:
                issues.append(f"{attachment.original_name} / {sheet.title}：未找到包含“净值日期、单位净值”的表头")
                continue
            header_index, mapping = matched
            for source_row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
                fund_code = code_value(cell_at(row, mapping.get("基金代码")))
                fund_name = text_value(cell_at(row, mapping.get("基金名称")))
                product_code = code_value(cell_at(row, mapping.get("产品代码")))
                product_name = text_value(cell_at(row, mapping.get("产品名称")))
                nav_date = parse_date(cell_at(row, mapping.get("净值日期")))
                unit_nav = parse_number(cell_at(row, mapping.get("单位净值")))
                cumulative_nav = parse_number(cell_at(row, mapping.get("累计净值")))

                # 空白尾行与无效记录都跳过；累计净值允许为空。
                has_identity = any((fund_code, fund_name, product_code, product_name))
                if not has_identity and nav_date is None and unit_nav is None:
                    continue
                if not has_identity or nav_date is None or unit_nav is None:
                    issues.append(
                        f"{attachment.original_name} / {sheet.title} / 第 {source_row_number} 行："
                        "缺少基金/产品标识、净值日期或单位净值，已跳过"
                    )
                    continue

                values: tuple[Any, ...] = (
                    fund_code,
                    fund_name,
                    product_code,
                    product_name,
                    nav_date,
                    unit_nav,
                    cumulative_nav,
                    attachment.original_name,
                    sheet.title,
                    source_row_number,
                )
                dedupe_key = (
                    (fund_code or fund_name).casefold(),
                    (product_code or product_name).casefold(),
                    nav_date.isoformat(),
                )
                extracted.append(ExtractedRow(values=values, dedupe_key=dedupe_key))
    finally:
        workbook.close()
    return extracted, issues


def deduplicate(rows: Iterable[ExtractedRow]) -> tuple[list[ExtractedRow], int, list[str]]:
    """相同主键保留第一条；净值不同则显式报告冲突，绝不静默覆盖。"""
    seen: dict[tuple[str, ...], ExtractedRow] = {}
    result: list[ExtractedRow] = []
    duplicate_count = 0
    conflicts: list[str] = []
    for row in rows:
        existing = seen.get(row.dedupe_key)
        if existing is not None:
            old_unit, old_cumulative = existing.values[5], existing.values[6]
            new_unit, new_cumulative = row.values[5], row.values[6]
            if old_unit == new_unit and old_cumulative == new_cumulative:
                duplicate_count += 1
            else:
                conflicts.append(
                    "同一基金/产品/日期出现不同净值，保留首次读取记录："
                    f"{existing.values[7]} 第 {existing.values[9]} 行（单位净值 {old_unit}）"
                    f"；跳过 {row.values[7]} 第 {row.values[9]} 行（单位净值 {new_unit}）"
                )
            continue
        seen[row.dedupe_key] = row
        result.append(row)
    return result, duplicate_count, conflicts


def make_sample_input(directory: Path) -> None:
    """生成两份固定、脱敏的 .xlsx 测试附件，供离线演示。"""
    directory.mkdir(parents=True, exist_ok=True)
    standard_path = directory / "01_standard_nav.xlsx"
    alias_path = directory / "02_alias_nav.xlsx"
    if standard_path.exists() or alias_path.exists():
        raise DemoError(f"示例目录已包含同名文件，请换一个空目录：{directory}")

    standard = Workbook()
    sheet = standard.active
    sheet.title = "净值数据"
    sheet.append(["基金代码", "基金名称", "产品代码", "产品名称", "净值日期", "单位净值", "累计净值"])
    sheet.append(["001234", "稳健增长A", "LB001", "稳健增长A产品", date(2026, 8, 28), 1.2345, 1.5678])
    sheet.append(["002345", "量化优选混合", "LB002", "量化优选产品", date(2026, 8, 29), 0.9876, 0.9876])
    sheet.append(["003456", "固收宝1号", "LB003", "固收宝1号产品", date(2026, 8, 29), 1.0820, 1.1200])
    standard.save(standard_path)
    standard.close()

    aliases = Workbook()
    sheet = aliases.active
    sheet.title = "估值数据"
    sheet.append(["示例：净值日报"])
    sheet.append([])
    sheet.append(["基金代码", "基金简称", "产品代码", "产品全称", "公布日期", "每份净值", "累计单位净值", "备注"])
    sheet.append(["001234", "稳健增长A", "LB001", "稳健增长A产品", "2026-08-28", 1.2345, 1.5678, "重复记录"])
    sheet.append(["004567", "新锐成长", "LB004", "新锐成长产品", "2026/08/30", 1.0560, 1.0560, "别名表头"])
    sheet.append(["005678", "灵活配置", "LB005", "灵活配置产品", "2026-08-30", 1.1030, None, "累计净值为空"])
    aliases.save(alias_path)
    aliases.close()


def write_output(rows: list[ExtractedRow], output_dir: Path, run_id: str) -> Path:
    """写出固定 10 列、单工作表的演示台账。"""
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"基金净值台账_{run_id}.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "净值台账"
    sheet.sheet_view.showGridLines = False
    sheet.append(list(OUTPUT_COLUMNS))
    for row in rows:
        sheet.append(list(row.values))

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_blue = Side(style="thin", color="B4C7E7")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin_blue)

    widths = (14, 20, 14, 24, 13, 12, 12, 28, 16, 12)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row_number in range(2, sheet.max_row + 1):
        sheet.cell(row_number, 5).number_format = "yyyy-mm-dd"
        sheet.cell(row_number, 6).number_format = "0.0000"
        sheet.cell(row_number, 7).number_format = "0.0000"
        sheet.cell(row_number, 10).alignment = Alignment(horizontal="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:J{sheet.max_row}"
    sheet.row_dimensions[1].height = 24
    workbook.save(output_path)
    workbook.close()
    return output_path


def positive_int(value: str) -> int:
    try:
        parsed = int(value)
    except ValueError as exc:
        raise argparse.ArgumentTypeError("请输入整数。") from exc
    if parsed < 0:
        raise argparse.ArgumentTypeError("不能小于 0。")
    return parsed


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="163 邮箱净值附件提取（演示版）。默认从 163 IMAP 抓取 .xlsx；--input-dir 可离线演示。"
    )
    parser.add_argument("--config", type=Path, default=APP_DIR / "config.env", help="配置文件路径")
    parser.add_argument("--input-dir", type=Path, help="本地 .xlsx 目录；指定后绝不连接邮箱")
    parser.add_argument(
        "--make-sample-input",
        type=Path,
        metavar="目录",
        help="生成两份脱敏的 .xlsx 示例附件，然后退出；不会连接邮箱",
    )
    parser.add_argument("--days", type=positive_int, help="覆盖配置中的检索天数")
    parser.add_argument("--max-rows", type=positive_int, help="输出最多多少条；0 表示不限制")
    parser.add_argument("--output-dir", type=Path, help="覆盖配置中的输出目录")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.make_sample_input:
        if args.input_dir:
            raise DemoError("--make-sample-input 不能与 --input-dir 同时使用。")
        make_sample_input(args.make_sample_input.resolve())
        ok(f"已生成离线示例附件：{args.make_sample_input.resolve()}")
        return 0

    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")

    if args.input_dir:
        attachments = attachments_from_local(args.input_dir.resolve())
        output_dir = (args.output_dir or APP_DIR / "output").resolve()
        max_rows = 5 if args.max_rows is None else args.max_rows
    else:
        config = load_env_file(args.config.resolve())
        try:
            configured_days = positive_int(config.get("LOOKBACK_DAYS", "7"))
            configured_max_rows = positive_int(config.get("MAX_ROWS", "5"))
            max_attachment_mb = positive_int(config.get("MAX_ATTACHMENT_MB", "20"))
        except argparse.ArgumentTypeError as exc:
            raise DemoError(f"config.env 配置值错误：{exc}") from exc
        days = configured_days if args.days is None else args.days
        max_rows = configured_max_rows if args.max_rows is None else args.max_rows
        output_dir = (
            args.output_dir.resolve()
            if args.output_dir
            else config_path_value(config.get("OUTPUT_DIR", ""), "output").resolve()
        )
        download_root = config_path_value(config.get("DOWNLOAD_DIR", ""), "downloads")
        subject_keyword = config.get("SUBJECT_KEYWORD", "").strip()
        attachments = fetch_from_163(
            config=config,
            run_directory=download_root / run_id,
            days=days,
            subject_keyword=subject_keyword,
            max_attachment_mb=max_attachment_mb or 20,
        )

    if not attachments:
        raise DemoError("没有找到可处理的 .xlsx 附件，本次未生成输出文件。")

    all_rows: list[ExtractedRow] = []
    all_issues: list[str] = []
    for attachment in attachments:
        info(f"解析附件：{attachment.original_name}")
        extracted, issues = extract_rows_from_attachment(attachment)
        all_rows.extend(extracted)
        all_issues.extend(issues)
        info(f"  提取到 {len(extracted)} 条有效记录。")

    normalized_rows, duplicate_count, conflicts = deduplicate(all_rows)
    if duplicate_count:
        warn(f"发现 {duplicate_count} 条完全重复记录，已按“首次读取优先”去重。")
    for conflict in conflicts:
        warn(conflict)
    for issue in all_issues:
        warn(issue)
    if not normalized_rows:
        raise DemoError("附件中没有满足固定格式要求的记录，本次未生成输出文件。")

    if max_rows:
        normalized_rows = normalized_rows[:max_rows]
    output_path = write_output(normalized_rows, output_dir, run_id)
    ok(f"已生成 {len(normalized_rows)} 条记录：{output_path}")
    info("固定输出字段：" + "、".join(OUTPUT_COLUMNS))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except DemoError as exc:
        print(f"[错误] {exc}", file=sys.stderr)
        raise SystemExit(2)
    except KeyboardInterrupt:
        print("\n[取消] 用户中断。", file=sys.stderr)
        raise SystemExit(130)
    except Exception:
        print("[错误] 程序发生未预期异常，详细信息如下：", file=sys.stderr)
        traceback.print_exc()
        raise SystemExit(1)
